from pyzabbix import ZabbixAPI
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from argparse import ArgumentParser
import configparser
import datetime
import sys
import urllib3
import requests

def make_timestamp(time):     return int(time.timestamp())
def make_datetime(timestamp): return datetime.datetime.fromtimestamp(timestamp)
def format_date(date):        return date.strftime('%d.%m %H:%M')
def parse_date(str):          return datetime.datetime.strptime(str, '%d.%m.%Y %H:%M')

# Read CMD arguments
parser = ArgumentParser()
parser.add_argument('date_till', type=str, help='Search till this date')
args = parser.parse_args()

try:
    date_till = parse_date(args.date_till)
except ValueError as e:
    print(e)
    print("Date must be in DD.MM.YYYY HH:MM format")
    sys.exit()

# Read config.ini for credentials
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')
login = config['credentials']['login']
password = config['credentials']['password']
zabbix_server_url = config['params']['zabbix_server_url']

# Inialize API access, disable SSL verification
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
s = requests.Session()
s.verify = False
zapi = ZabbixAPI(zabbix_server_url, s)
zapi.login(login, password)

hosts = zapi.host.get(#groupids=group_ids,
                      monitored_hosts=True,
                      output=['hostid', 'host'])
print(f"Monitored hosts: {len(hosts)}")

# Abort if no hosts were found
if len(hosts) == 0:
    print("No hosts were founds")
    sys.exit()

# For every host get corresponding itemid for its software list
host_ids = [ h['hostid'] for h in hosts ]
items = zapi.item.get(hostids=host_ids,
                      output=['hostid', 'lastclock', 'key_'],
                      sortfield='itemid',
                      monitored=True,
                      filter={ 'key_' : ['ubuntu.soft', 'system.sw.packages'] },
                      selectHosts=['host'])
print(f"Enabled items: {len(items)}")

# Filtering out items if there are more than one for a single host
filtered_items = []
for host in hosts:
    host_items = [item for item in items if item['hostid'] == host['hostid']]
    if not host_items:
        continue
    elif len(host_items) > 1:
        host_item = next(item for item in host_items if item['key_'] == 'ubuntu.soft')
    else:
        host_item = host_items[0]
    filtered_items.append(host_item)
items = filtered_items
print(f"Items with unique host: {len(items)}")

# For every itemid get its value history
date_from = date_till - datetime.timedelta(hours=18 + 1)
print(f"Searching for history from {format_date(date_from)} to {format_date(date_till)}")
item_ids = [ i['itemid'] for i in items ]
history = zapi.history.get(itemids=item_ids,
                           history=4,
                           sortfield='clock',
                           sortorder='DESC',
                           time_from=make_timestamp(date_from),
                           time_till=make_timestamp(date_till),
                           output=['itemid', 'clock', 'value'])
if not history:
    print('No history entries found')
    sys.exit()
print(f"History length: {len(history)}")

# We have all the data, now to combine it together
for host in hosts:
    item = next((item for item in items if item['hostid'] == host['hostid']), None)
    if not item:
        continue
    host['itemid'] = host_item['itemid']
    package_lists = [h['value'] for h in history if h['itemid'] == item['itemid']]
    if package_lists:
        newest_package_list = package_lists[0]
        oldest_package_list = package_lists[-1]
        if item['key_'] == 'ubuntu.soft':
            host['new_packages'] = set(newest_package_list.split('\n'))
            host['old_packages'] = set(oldest_package_list.split('\n'))
        else: # Assuming centOS alsways have [...] in front
            start = newest_package_list.find(']') + 2
            host['new_packages'] = set(newest_package_list[start:].split(', ') )
            host['old_packages'] = set(oldest_package_list[start:].split(', ') )
hosts.sort(key=lambda h: h['host'])

# Compile lists of new and removed software via set differences
for host in hosts:
    if 'new_packages' in host:
        installed = list(host['new_packages'] - host['old_packages'])
        removed   = list(host['old_packages'] - host['new_packages'])
        host['installed'] = sorted(installed)
        host['removed']   = sorted(removed)
        host['newest_software_lsit'] = sorted(list(host['new_packages']))
    else:
        host['installed'] = ['No Data']
        host['removed']   = ['No Data']
        host['newest_software_lsit'] = ['No Data']

# Compiling a dictionary to group up hosts with identical changes
host_groups = {}
for host in hosts:
    key_tuple = tuple(host['newest_software_lsit'])
    # Do not include hosts with no history or no changes in the output
    # if not key_tuple or key_tuple[0] == 'No Data':
    #     continue
    if not key_tuple in host_groups.keys():
        host_groups[key_tuple] = []
    host_groups[key_tuple].append(host)
print(f"Total groups of hosts with same software: {len(host_groups)}")

# Output to a .txt file
def output_txt(filename):
    with open(filename, 'w') as f:
        for key_tuple, hosts in host_groups.items():
            for host in hosts:
                print(host['host'], file=f)
            host = hosts[0]
            if host['newest_software_lsit']:
                print("\nPackages:", file=f)
                for new_package in host['newest_software_lsit']:
                    print(new_package, file=f, end='')
                print('', file=f)
            print('----------------------------------------------------------------------------------------', file=f)
output_txt('report.txt')

# Output to an .xlsx table
def output_xlsx(filename):
    def set_border(ws, cell_range, bottom_row, top_row):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                if cell.row == bottom_row:
                    cell.border += Border(top=thin)
                if cell.row == top_row:
                    cell.border += Border(bottom=thin)
                cell.border += Border(left=thin, right=thin)

    workbook = Workbook()
    sheet = workbook.active

    top_row = 1
    for key_tuple, hosts in host_groups.items():
        # sheet.cell(row=top_row, column=1).value = 'Хосты'
        # sheet.cell(row=top_row, column=2).value = 'Пакеты'
        row = top_row
        for host in hosts:
            sheet.cell(row=row, column=1).value = host['host']
            row += 1

        # row = top_row + 1
        # for package in hosts[0]['newest_software_lsit']:
        #     sheet.cell(row=row, column=2).value = package
        #     row += 1

        package_list = ''
        for package in hosts[0]['newest_software_lsit']:
            package_list += package + ', '
        package_list = package_list[:-2]
        sheet.cell(row=top_row, column=2).value = package_list

        old_top_row = top_row
        top_row += max(len(hosts), 1)
        set_border(sheet, f"A{old_top_row}:B{top_row - 1}", old_top_row, top_row - 1)

    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) 
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 5
    workbook.save(filename=filename)
output_xlsx('report.xlsx')